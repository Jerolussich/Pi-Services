#!/usr/bin/env python3
import argparse
import json
import os
import sqlite3
import sys
import traceback
from datetime import date, datetime, timedelta
from calendar import monthrange

import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TOKENS_FILE = os.path.join(os.path.dirname(__file__), "tokens.json")
EXPORTS_DIR = os.path.join(os.path.dirname(__file__), "exports")
DB_PATH     = os.path.join(EXPORTS_DIR, "fitbit.db")
BASE_URL = "https://api.fitbit.com"

def load_tokens():
    with open(TOKENS_FILE) as f:
        return json.load(f)

def refresh_access_token():
    t = load_tokens()
    response = requests.post(
        "https://api.fitbit.com/oauth2/token",
        data={
            "grant_type": "refresh_token",
            "refresh_token": t["refresh_token"],
        },
        auth=(t["client_id"], t["client_secret"]),
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )
    if response.status_code == 200:
        new_tokens = response.json()
        t["access_token"] = new_tokens["access_token"]
        t["refresh_token"] = new_tokens["refresh_token"]
        with open(TOKENS_FILE, "w") as f:
            json.dump(t, f, indent=2)
        print("✓ Token renovado automáticamente.")
        return t
    else:
        raise Exception(f"No se pudo renovar el token: {response.status_code} {response.text}")

def api_get(token, path):
    """Hace un GET a la API de Fitbit con manejo de errores"""
    headers = {"Authorization": f"Bearer {token}"}
    url = f"{BASE_URL}{path}" if path.startswith("/") else path
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        raise Exception(f"{response.status_code}: {response.text[:200]}")

def get_month_range(year, month):
    first = date(year, month, 1)
    last = date(year, month, monthrange(year, month)[1])
    return first, last

def date_range(start, end):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)

def fetch_steps(token, start, end):
    rows = []
    s, e = start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")
    data     = api_get(token, f"/1/user/-/activities/steps/date/{s}/{e}.json")
    calories = api_get(token, f"/1/user/-/activities/calories/date/{s}/{e}.json")
    distance = api_get(token, f"/1/user/-/activities/distance/date/{s}/{e}.json")
    active   = api_get(token, f"/1/user/-/activities/minutesVeryActive/date/{s}/{e}.json")

    steps_map  = {d["dateTime"]: d["value"] for d in data["activities-steps"]}
    cal_map    = {d["dateTime"]: d["value"] for d in calories["activities-calories"]}
    dist_map   = {d["dateTime"]: d["value"] for d in distance["activities-distance"]}
    active_map = {d["dateTime"]: d["value"] for d in active["activities-minutesVeryActive"]}

    for d in date_range(start, end):
        ds = d.strftime("%Y-%m-%d")
        rows.append({
            "Fecha": ds,
            "Pasos": steps_map.get(ds, 0),
            "Calorías": cal_map.get(ds, 0),
            "Distancia (km)": dist_map.get(ds, 0),
            "Minutos activos": active_map.get(ds, 0),
        })
    return rows

def fetch_sleep(token, start, end):
    rows = []
    for d in date_range(start, end):
        try:
            data = api_get(token, f"/1.2/user/-/sleep/date/{d.strftime('%Y-%m-%d')}.json")
            summary = data.get("summary", {})
            stages = summary.get("stages", {})
            rows.append({
                "Fecha": d.strftime("%Y-%m-%d"),
                "Total (min)": summary.get("totalMinutesAsleep", 0),
                "En cama (min)": summary.get("totalTimeInBed", 0),
                "Eficiencia (%)": data.get("sleep", [{}])[0].get("efficiency", 0) if data.get("sleep") else 0,
                "Deep (min)": stages.get("deep", 0),
                "Light (min)": stages.get("light", 0),
                "REM (min)": stages.get("rem", 0),
                "Despierto (min)": stages.get("wake", 0),
            })
        except Exception as e:
            print(f"  Sleep {d}: {e}")
            rows.append({"Fecha": d.strftime("%Y-%m-%d")})
    return rows

def fetch_heart_rate(token, start, end):
    rows = []
    s, e = start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")
    data = api_get(token, f"/1/user/-/activities/heart/date/{s}/{e}.json")
    for entry in data["activities-heart"]:
        val = entry.get("value", {})
        zones = {z["name"]: z.get("minutes", 0) for z in val.get("heartRateZones", [])}
        rows.append({
            "Fecha": entry["dateTime"],
            "Resting HR (bpm)": val.get("restingHeartRate", ""),
            "Out of Range (min)": zones.get("Out of Range", 0),
            "Fat Burn (min)": zones.get("Fat Burn", 0),
            "Cardio (min)": zones.get("Cardio", 0),
            "Peak (min)": zones.get("Peak", 0),
        })
    return rows

def fetch_spo2(token, start, end):
    rows = []
    s, e = start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")
    try:
        data = api_get(token, f"/1/user/-/spo2/date/{s}/{e}.json")
        for entry in data:
            rows.append({
                "Fecha": entry["dateTime"],
                "SpO2 promedio (%)": entry.get("value", {}).get("avg", ""),
                "SpO2 min (%)": entry.get("value", {}).get("min", ""),
                "SpO2 max (%)": entry.get("value", {}).get("max", ""),
            })
    except Exception as e:
        print(f"  SpO2 error: {e}")
    return rows

def fetch_activity_logs(token, start, end):
    rows = []
    current = start
    while current <= end:
        try:
            data = api_get(token,
                f"/1/user/-/activities/list.json?afterDate={current.strftime('%Y-%m-%d')}&sort=asc&limit=20&offset=0"
            )
            activities = data.get("activities", [])
            for activity in activities:
                start_time = activity.get("startTime", "")
                rows.append({
                    "Fecha": start_time[:10] if start_time else current.strftime("%Y-%m-%d"),
                    "Hora inicio": start_time[11:16] if start_time else "",
                    "Actividad": activity.get("activityName", ""),
                    "Duración (min)": round(activity.get("duration", 0) / 60000),
                    "Calorías": activity.get("calories", ""),
                    "Distancia (km)": round(activity.get("distance", 0), 2) if activity.get("distance") else "",
                    "Pasos": activity.get("steps", ""),
                    "Frecuencia cardíaca avg": activity.get("averageHeartRate", ""),
                })
            if activities:
                last = activities[-1].get("startTime", "")[:10]
                current = date.fromisoformat(last) + timedelta(days=1)
            else:
                current += timedelta(days=1)
        except Exception as e:
            print(f"  Activity logs {current}: {e}")
            current += timedelta(days=1)
    rows = [r for r in rows if start.strftime("%Y-%m-%d") <= r["Fecha"] <= end.strftime("%Y-%m-%d")]
    return rows

def month_already_exported(ws, year, month):
    prefix = f"{year}-{month:02d}"
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0]).startswith(prefix):
            return True
    return False

def write_sheet(wb, title, rows, headers, year=None, month=None):
    if title in wb.sheetnames:
        ws = wb[title]
        if year and month and month_already_exported(ws, year, month):
            print(f"  ⚠️  {title}: {year}-{month:02d} ya exportado, saltando...")
            return ws
        next_row = ws.max_row + 1
    else:
        ws = wb.create_sheet(title=title)
        header_fill = PatternFill("solid", fgColor="2D6A9F")
        header_font = Font(bold=True, color="FFFFFF")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = max(len(header) + 4, 14)
        next_row = 2

    for r, row in enumerate(rows, next_row):
        for c, key in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=row.get(key, ""))

    ws.auto_filter.ref = ws.dimensions
    return ws

def save_to_sqlite(data_dict, db_path):
    """Guarda todos los datos en SQLite"""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Actividad
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS actividad (
            fecha TEXT PRIMARY KEY,
            pasos INTEGER,
            calorias INTEGER,
            distancia_km REAL,
            minutos_activos INTEGER
        )
    """)
    for row in data_dict.get("steps", []):
        cursor.execute("""
            INSERT OR REPLACE INTO actividad VALUES (?,?,?,?,?)
        """, (row.get("Fecha"), row.get("Pasos"), row.get("Calorías"),
              row.get("Distancia (km)"), row.get("Minutos activos")))

    # Sueño
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS sueno (
            fecha TEXT PRIMARY KEY,
            total_min INTEGER,
            en_cama_min INTEGER,
            eficiencia_pct REAL,
            deep_min INTEGER,
            light_min INTEGER,
            rem_min INTEGER,
            despierto_min INTEGER
        )
    """)
    for row in data_dict.get("sleep", []):
        cursor.execute("""
            INSERT OR REPLACE INTO sueno VALUES (?,?,?,?,?,?,?,?)
        """, (row.get("Fecha"), row.get("Total (min)"), row.get("En cama (min)"),
              row.get("Eficiencia (%)"), row.get("Deep (min)"), row.get("Light (min)"),
              row.get("REM (min)"), row.get("Despierto (min)")))

    # Heart Rate
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS heart_rate (
            fecha TEXT PRIMARY KEY,
            resting_hr INTEGER,
            out_of_range_min INTEGER,
            fat_burn_min INTEGER,
            cardio_min INTEGER,
            peak_min INTEGER
        )
    """)
    for row in data_dict.get("hr", []):
        cursor.execute("""
            INSERT OR REPLACE INTO heart_rate VALUES (?,?,?,?,?,?)
        """, (row.get("Fecha"), row.get("Resting HR (bpm)"), row.get("Out of Range (min)"),
              row.get("Fat Burn (min)"), row.get("Cardio (min)"), row.get("Peak (min)")))

    # Ejercicios
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ejercicios (
            fecha TEXT,
            hora_inicio TEXT,
            actividad TEXT,
            duracion_min INTEGER,
            calorias INTEGER,
            distancia_km REAL,
            pasos INTEGER,
            hr_avg INTEGER,
            PRIMARY KEY (fecha, hora_inicio)
        )
    """)
    for row in data_dict.get("act_logs", []):
        cursor.execute("""
            INSERT OR REPLACE INTO ejercicios VALUES (?,?,?,?,?,?,?,?)
        """, (row.get("Fecha"), row.get("Hora inicio"), row.get("Actividad"),
              row.get("Duración (min)"), row.get("Calorías"), row.get("Distancia (km)"),
              row.get("Pasos"), row.get("Frecuencia cardíaca avg")))

    conn.commit()
    conn.close()
    print("✓ Datos guardados en SQLite.")

def previous_month():
    today = date.today()
    if today.month == 1:
        return today.year - 1, 12
    return today.year, today.month - 1


def ensure_ingest_runs_table(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ingest_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            started_at TEXT NOT NULL,
            finished_at TEXT,
            status TEXT NOT NULL,
            error TEXT,
            source TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_ingest_runs_ym
            ON ingest_runs(year, month, started_at DESC)
    """)


def start_ingest_run(year, month, source):
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    try:
        ensure_ingest_runs_table(conn)
        cur = conn.execute(
            "INSERT INTO ingest_runs (year, month, started_at, status, source) "
            "VALUES (?, ?, ?, 'running', ?)",
            (year, month, datetime.utcnow().isoformat(timespec="seconds"), source),
        )
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def finish_ingest_run(run_id, status, error=None):
    if run_id is None:
        return
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute(
            "UPDATE ingest_runs SET finished_at=?, status=?, error=? WHERE id=?",
            (datetime.utcnow().isoformat(timespec="seconds"), status, error, run_id),
        )
        conn.commit()
    finally:
        conn.close()


def run(year, month, source):
    start, end = get_month_range(year, month)
    filename = os.path.join(EXPORTS_DIR, "fitbit_data.xlsx")

    print(f"Exportando datos de {start} a {end} (source={source})...")

    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        print("Appendando a fitbit_data.xlsx existente...")
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        print("Creando fitbit_data.xlsx nuevo...")

    t = refresh_access_token()
    token = t["access_token"]

    print("Obteniendo steps y actividad...")
    steps = fetch_steps(token, start, end)
    write_sheet(wb, "Actividad", steps, ["Fecha", "Pasos", "Calorías", "Distancia (km)", "Minutos activos"], year, month)

    print("Obteniendo sueño...")
    sleep = fetch_sleep(token, start, end)
    write_sheet(wb, "Sueño", sleep, ["Fecha", "Total (min)", "En cama (min)", "Eficiencia (%)", "Deep (min)", "Light (min)", "REM (min)", "Despierto (min)"], year, month)

    print("Obteniendo frecuencia cardíaca...")
    hr = fetch_heart_rate(token, start, end)
    write_sheet(wb, "Heart Rate", hr, ["Fecha", "Resting HR (bpm)", "Out of Range (min)", "Fat Burn (min)", "Cardio (min)", "Peak (min)"], year, month)

    print("Obteniendo SpO2...")
    spo2 = fetch_spo2(token, start, end)
    write_sheet(wb, "SpO2", spo2, ["Fecha", "SpO2 promedio (%)", "SpO2 min (%)", "SpO2 max (%)"], year, month)

    print("Obteniendo logs de actividades...")
    act_logs = fetch_activity_logs(token, start, end)
    write_sheet(wb, "Ejercicios", act_logs, ["Fecha", "Hora inicio", "Actividad", "Duración (min)", "Calorías", "Distancia (km)", "Pasos", "Frecuencia cardíaca avg"], year, month)

    wb.save(filename)

    save_to_sqlite({
        "steps": steps,
        "sleep": sleep,
        "hr": hr,
        "act_logs": act_logs,
    }, DB_PATH)

    print(f"\n✓ Exportado: {filename}")


def parse_args():
    p = argparse.ArgumentParser(description="Export Fitbit data for a given month.")
    p.add_argument("--year", type=int, help="Year (YYYY). Defaults to previous month's year.")
    p.add_argument("--month", type=int, choices=range(1, 13), metavar="MM",
                   help="Month (1-12). Defaults to previous month.")
    p.add_argument("--source", choices=("manual", "scheduled"), default="scheduled",
                   help="Who triggered this run (recorded in ingest_runs).")
    args = p.parse_args()
    if (args.year is None) != (args.month is None):
        p.error("--year and --month must be given together")
    return args


def main():
    args = parse_args()
    if args.year is None:
        year, month = previous_month()
    else:
        year, month = args.year, args.month

    run_id = start_ingest_run(year, month, args.source)
    try:
        run(year, month, args.source)
    except Exception as e:
        tb = traceback.format_exc()
        print(tb, file=sys.stderr)
        finish_ingest_run(run_id, "error", tb[-2000:])
        sys.exit(1)
    else:
        finish_ingest_run(run_id, "success")


if __name__ == "__main__":
    main()
